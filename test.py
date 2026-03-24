from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Data Dictionary"

# ── Styles ──────────────────────────────────────────────────────────────────
hdr_fill   = PatternFill("solid", start_color="1F3864")   # dark navy
sec_fill   = PatternFill("solid", start_color="2E75B6")   # medium blue
alt_fill   = PatternFill("solid", start_color="D6E4F0")   # light blue
white_fill = PatternFill("solid", start_color="FFFFFF")
global_fill= PatternFill("solid", start_color="E2EFDA")   # light green for globals
param_fill = PatternFill("solid", start_color="FFF2CC")   # yellow for params
ret_fill   = PatternFill("solid", start_color="FCE4D6")   # orange for returns
cls_fill   = PatternFill("solid", start_color="EAD1DC")   # pink for class attrs

hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
sec_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
body_font  = Font(name="Arial", size=9)
bold_font  = Font(name="Arial", bold=True, size=9)

thin = Side(style="thin", color="B8CCE4")
thick = Side(style="medium", color="2E75B6")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top", wrap_text=True)

def style_header(cell, text):
    cell.value = text
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)

def style_section(ws, row, label, fill=sec_fill):
    ws.merge_cells(f"A{row}:I{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = sec_font
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(left=thick, right=thick, top=thin, bottom=thin)

def add_row(ws, row, data, fill=None):
    fills = [white_fill, alt_fill]
    row_fill = fill if fill else fills[row % 2]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = body_font
        c.fill = row_fill
        c.alignment = wrap
        c.border = cell_border

# ── Column Headers ───────────────────────────────────────────────────────────
headers = ["Variable", "Data Type", "Format for Display", "Size (bytes)",
           "Size for Display", "Description", "Example", "Validation", "Scope / Source"]

for col, h in enumerate(headers, 1):
    style_header(ws.cell(row=1, column=col), h)

# ── Column widths ────────────────────────────────────────────────────────────
widths = [28, 18, 20, 13, 15, 48, 32, 36, 26]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

ROW = 2  # current write row

# ───────────────────────────────────────────────────────────────────────────
# SECTION: GLOBAL VARIABLES (main.py)
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  GLOBAL VARIABLES — main.py"); ROW += 1

globals_data = [
    ["NumberOfMines",           "int",    "Integer",              "28",  "≤10 chars",  "Number of mines owned by the player's colony. Controls ore production and population-to-mine ratio checks.",            "8",           "Must be ≥ 1 during gameplay. Ratio Population/NumberOfMines ≥ 10 required.",    "Global – main.py"],
    ["Population",              "int",    "Integer",              "28",  "≤10 chars",  "Current number of people living in the colony. Affects satisfaction calculation, ore production, and loss conditions.","120",          "Must be > 0. Loss triggered if < 30.",                                        "Global – main.py"],
    ["Money",                   "int/float","Integer or decimal", "28",  "≤12 chars",  "Player's current currency balance. Adjusted by buying/selling mines, ore, and food each term.",                       "3500",         "No lower bound enforced; negative balance possible.",                         "Global – main.py"],
    ["FoodPrice",               "int",    "Integer ($ per unit)", "28",  "≤6 chars",   "Market price per unit of food this term. Randomised at game start.",                                                  "32",           "Random range: 21–40.",                                                        "Global – main.py"],
    ["OreProducion",            "int",    "Integer (tons/mine)",  "28",  "≤6 chars",   "Tons of ore each mine produces per term. Fluctuates with satisfaction. Note: typo in source (missing 't').",          "90",           "Minimum enforced at 65 per UpdateVariables().",                               "Global – main.py"],
    ["OrePrice",                "int",    "Integer ($ per ton)",  "28",  "≤6 chars",   "Market price per ton of ore. Randomised each term.",                                                                   "14",           "Random range: 8–19.",                                                         "Global – main.py"],
    ["MinePrice",               "int",    "Integer ($)",          "28",  "≤8 chars",   "Market price to buy/sell a mine. Randomised each term.",                                                              "3200",         "Random range: 2001–4000.",                                                    "Global – main.py"],
    ["currentYear",             "int",    "Integer",              "28",  "≤4 chars",   "Current game term (year). Starts at 0, incremented by UpdateVariables() each turn.",                                   "3",            "0 to yearsToSurvive+1.",                                                      "Global – main.py"],
    ["oreInStorage",            "int/float","Integer",            "28",  "≤8 chars",   "Cumulative ore in storage, not yet sold. Increases by OreProducion×NumberOfMines per term.",                           "450",          "≥ 0.",                                                                        "Global – main.py"],
    ["currentSatfication",      "float",  "Decimal (2 d.p.)",     "24",  "≤5 chars",   "Colony satisfaction score. Affects ore production growth and population change. Typo in source (missing 'is').",      "0.95",         "Loss if < 0.6. Value can exceed 1.0 with high food purchases.",               "Global – main.py"],
    ["running",                 "bool",   "True / False",         "28",  "5 chars",    "Main game loop flag. Set to False by QuitGame() to exit the application.",                                             "True",         "Boolean only.",                                                               "Global – main.py"],
    ["yearsToSurvive",          "int",    "Integer",              "28",  "≤2 chars",   "Number of terms the player must survive. Set by difficulty buttons.",                                                   "10",           "5 (Easy), 10 (Normal), 20 (Hard).",                                           "Global – main.py"],
    ["HARD_MODE_YEARS_TO_SURVIVE",  "int","Integer (constant)",   "28",  "2 chars",    "Constant: number of terms for Hard difficulty.",                                                                        "20",           "Fixed at 20.",                                                                "Constant – main.py"],
    ["NORMAL_MODE_YEARS_TO_SURVIVE","int","Integer (constant)",   "28",  "2 chars",    "Constant: number of terms for Normal difficulty.",                                                                      "10",           "Fixed at 10.",                                                                "Constant – main.py"],
    ["EASY_MODE_YEARS_TO_SURVIVE",  "int","Integer (constant)",   "28",  "1 char",     "Constant: number of terms for Easy difficulty.",                                                                        "5",            "Fixed at 5.",                                                                 "Constant – main.py"],
    ["DEFAULT_SCREEN",          "str",    "File path string",     "variable","path string","Raw string path to the default background image asset.",                                                            r"Assets\Background\SpaceMinesBGv3.png","Must be a valid relative file path to a .png image.",  "Constant – main.py"],
    ["OVERWORK_LOSS_SCREEN",    "str",    "File path string",     "variable","path string","Path to the overworked-population loss screen image.",                                                              r"Assets\Background\LoseScreens\OverworkedPopulationSpaceColony.png","Valid .png path.","Constant – main.py"],
    ["REVOLT_LOSS_SCREEN",      "str",    "File path string",     "variable","path string","Path to the revolt loss screen image.",                                                                             r"Assets\Background\LoseScreens\SpaceColonyRevolt.png","Valid .png path.",              "Constant – main.py"],
    ["NOT_ENOUGH_PEOPLE_SCREEN","str",    "File path string",     "variable","path string","Path to the not-enough-people loss screen image.",                                                                  r"Assets\Background\LoseScreens\NotEnoughPeopleSpaceColony.png","Valid .png path.",     "Constant – main.py"],
]
for i, r in enumerate(globals_data):
    add_row(ws, ROW, r + [], global_fill if i % 2 == 0 else PatternFill("solid", start_color="C6EFCE"))
    ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: GLOBAL VARIABLES (Display.py)
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  GLOBAL VARIABLES — Display.py"); ROW += 1
display_globals = [
    ["ASPECT_RATIO",    "float",    "Decimal",              "24",  "6 chars",    "Fixed 16:9 aspect ratio used to constrain window resizing.",                   "1.7778",       "Fixed at 16/9.",                          "Global – Display.py"],
    ["BACKGROUND_COLOR","tuple",    "(R, G, B) integers",   "72",  "(0,0,255)",  "RGB colour tuple for screen background fill. Currently pure blue.",            "(0, 0, 255)",  "Each channel 0–255.",                     "Global – Display.py"],
    ["screen",          "pygame.Surface","N/A (object)",    "variable","N/A",    "The pygame display surface. Re-assigned on window resize events.",              "pygame.Surface object","Must be a valid pygame Surface.",   "Global – Display.py"],
    ["running",         "bool",     "True / False",         "28",  "5 chars",    "Duplicate loop flag also present in main.py. Controls main game loop.",         "True",         "Boolean only.",                           "Global – Display.py"],
]
for i, r in enumerate(display_globals):
    add_row(ws, ROW, r, global_fill if i % 2 == 0 else PatternFill("solid", start_color="C6EFCE"))
    ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: GLOBAL VARIABLES (Button.py)
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  GLOBAL VARIABLES — Button.py"); ROW += 1
btn_globals = [
    ["buttonList",  "dict",  "{ 'MouseUp': [...], 'MouseDown': [...] }","variable","N/A","Module-level dictionary holding lists of all registered mouse event callbacks for all Button instances.","{\"MouseUp\":[fn1,fn2], \"MouseDown\":[fn3]}","Keys must be 'MouseUp' and 'MouseDown'; values are lists of callables.","Global – Button.py"],
]
add_row(ws, ROW, btn_globals[0], global_fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: GLOBAL VARIABLES (GUIBase.py)
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  GLOBAL VARIABLES — GUIBase.py"); ROW += 1
gui_globals = [
    ["guiAssetList","list","List of GUIBase objects","variable","N/A","Module-level list of every GUIBase instance created. Iterated each frame by GetGuiAssets() to render the scene.","[<GUIBase>, <TextLabel>, ...]","Must only contain GUIBase (or subclass) instances.","Global – GUIBase.py"],
]
add_row(ws, ROW, gui_globals[0], global_fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: GLOBAL VARIABLES (TextLabel.py)
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  GLOBAL VARIABLES — TextLabel.py"); ROW += 1
tl_globals = [
    ["fontCache",   "dict", "{ font_name: SysFont }",  "variable","N/A","Module-level font cache to avoid re-loading the same font repeatedly. Keyed by font name string.",                 "{ 'monospace': <SysFont> }","Values must be pygame.freetype.SysFont objects.","Global – TextLabel.py"],
    ["seperator",   "str",  "Single character",        "50",      "1 char","Word separator used when splitting/joining text for wrapping. Note: typo in source ('sepErator').",            "\" \"",            "Single whitespace character.",                   "Global – TextLabel.py"],
]
for i, r in enumerate(tl_globals):
    add_row(ws, ROW, r, global_fill if i % 2 == 0 else PatternFill("solid", start_color="C6EFCE"))
    ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: GLOBAL VARIABLES (Image.py)
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  GLOBAL VARIABLES — Image.py"); ROW += 1
img_globals = [
    ["CachedImages","dict","{ path_str: Surface }","variable","N/A","Module-level image cache. Stores loaded pygame Surfaces keyed by file path to prevent redundant disk reads.","{ r'Assets\\bg.png': <Surface> }","Keys must be valid file path strings; values must be pygame.Surface.","Global – Image.py"],
]
add_row(ws, ROW, img_globals[0], global_fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: FUNCTIONS – GeneralFunctions.py
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  FUNCTIONS — GeneralFunctions.py"); ROW += 1

style_section(ws, ROW, "   CreateUniqueKeyForMap(map)", PatternFill("solid", start_color="4472C4")); ROW += 1
gf_rows = [
    ["map",     "dict | None", "N/A (object)",     "variable","N/A","PARAMETER. The dictionary to generate a unique key for. If None the function returns early without generating a key.","{}","Must be a dict or None.",            "Param – CreateUniqueKeyForMap"],
    ["key",     "uuid.UUID",   "UUID string",      "16",      "36 chars","LOCAL. UUID generated to use as a dictionary key. Re-generated in a while loop until it is not already present in map.","uuid4() object","Must not already exist in map.",  "Local – CreateUniqueKeyForMap"],
    ["key",     "uuid.UUID | None","UUID string or None","16","36 chars / None","RETURN VALUE. The unique UUID key, or None (implicit) if map was None.",                                    "uuid.UUID('3f2...')","None if map parameter was None; otherwise a UUID not present in map.", "Return – CreateUniqueKeyForMap"],
]
for i, r in enumerate(gf_rows):
    add_row(ws, ROW, r, param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill))
    ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: FUNCTIONS – SpatialQueryLibrary.py
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  FUNCTIONS — SpatialQueryLibrary.py"); ROW += 1
style_section(ws, ROW, "   PointInsideRectange(RecSize, RecPos, MousePos)  [Note: typo in source – 'Rectange']", PatternFill("solid", start_color="4472C4")); ROW += 1
sq_rows = [
    ["RecSize",  "tuple",  "(float, float)", "72",  "(w, h)",       "PARAMETER. Width and height of the rectangle as normalised fractions of screen dimensions.",                "(0.1, 0.05)",  "Must be a 2-tuple of non-negative floats.",                           "Param – PointInsideRectange"],
    ["RecPos",   "tuple",  "(float, float)", "72",  "(x, y)",       "PARAMETER. Centre position of the rectangle as normalised fractions of screen dimensions.",                "(0.5, 0.5)",   "Must be a 2-tuple of floats in range 0.0–1.0.",                       "Param – PointInsideRectange"],
    ["MousePos", "tuple",  "(float, float)", "72",  "(x, y)",       "PARAMETER. Mouse cursor position as normalised fractions of screen dimensions.",                           "(0.48, 0.51)", "Must be a 2-tuple of non-negative floats.",                           "Param – PointInsideRectange"],
    ["x_pos",   "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Unpacked x-coordinate of the rectangle centre.",                                                     "0.5",          "Derived from RecPos[0].",                                             "Local – PointInsideRectange"],
    ["y_pos",   "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Unpacked y-coordinate of the rectangle centre.",                                                     "0.5",          "Derived from RecPos[1].",                                             "Local – PointInsideRectange"],
    ["x_size",  "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Unpacked width of the rectangle.",                                                                   "0.1",          "Derived from RecSize[0].",                                            "Local – PointInsideRectange"],
    ["y_size",  "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Unpacked height of the rectangle.",                                                                  "0.05",         "Derived from RecSize[1].",                                            "Local – PointInsideRectange"],
    ["xMP",     "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Unpacked x-coordinate of mouse position.",                                                           "0.48",         "Derived from MousePos[0].",                                           "Local – PointInsideRectange"],
    ["yMP",     "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Unpacked y-coordinate of mouse position.",                                                           "0.51",         "Derived from MousePos[1].",                                           "Local – PointInsideRectange"],
    ["xlb",     "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Left bound of rectangle: x_pos − x_size/2.",                                                        "0.45",         "Must be < xrb.",                                                      "Local – PointInsideRectange"],
    ["xrb",     "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Right bound of rectangle: x_pos + x_size/2.",                                                       "0.55",         "Must be > xlb.",                                                      "Local – PointInsideRectange"],
    ["ytb",     "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Top bound of rectangle: y_pos − y_size/2.",                                                         "0.475",        "Must be < ybb.",                                                      "Local – PointInsideRectange"],
    ["ybb",     "float",  "Decimal",         "24",  "N/A (local)",  "LOCAL. Bottom bound of rectangle: y_pos + y_size/2.",                                                      "0.525",        "Must be > ytb.",                                                      "Local – PointInsideRectange"],
    ["(return)","bool",   "True / False",    "28",  "5 chars",      "RETURN VALUE. True if mouse is inside the rectangle; False otherwise.",                                    "True",         "Boolean only.",                                                       "Return – PointInsideRectange"],
]
for r in sq_rows:
    fill = param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill)
    add_row(ws, ROW, r, fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: FUNCTIONS – Display.py
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  FUNCTIONS — Display.py"); ROW += 1
style_section(ws, ROW, "   TickDisplay()", PatternFill("solid", start_color="4472C4")); ROW += 1
td_rows = [
    ["screen",    "pygame.Surface","N/A (object)","variable","N/A","LOCAL (via global). The global display surface used to fill the background and draw all GUI objects.",  "pygame.Surface","Must be initialised pygame Surface.",  "Global (used) – Display.py"],
    ["UIAssets",  "list",           "List of GUIBase","variable","N/A","LOCAL. Sorted list of GUI objects returned by GetGuiAssets(). Iterated to call refresh() on each.","[<GUIBase>, ...]","Must be iterable; elements must have .refresh() method.","Local – TickDisplay"],
    ["guiObject", "GUIBase subclass","N/A (object)","variable","N/A","LOCAL. Loop variable; each GUI element in UIAssets is refreshed in turn.",                          "<TextLabel>","Must be a GUIBase subclass instance.",     "Local – TickDisplay"],
    ["(return)",  "None",           "None",        "N/A",     "N/A","RETURN VALUE. TickDisplay() returns nothing; side effect is screen redraw.",                        "None",         "Always None.",                             "Return – TickDisplay"],
]
for r in td_rows:
    fill = param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill)
    add_row(ws, ROW, r, fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: FUNCTIONS – GUIBase.py
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  FUNCTIONS — GUIBase.py"); ROW += 1
style_section(ws, ROW, "   clamp(value, minVal, maxVal)", PatternFill("solid", start_color="4472C4")); ROW += 1
clamp_rows = [
    ["value",    "int | float","Numeric","variable","N/A","PARAMETER. The value to be clamped.",                            "1.2",      "Must be numeric (int or float).",                 "Param – clamp"],
    ["minVal",   "int | float","Numeric","variable","N/A","PARAMETER. Lower bound of the clamp range.",                     "0",        "Must be ≤ maxVal.",                               "Param – clamp"],
    ["maxVal",   "int | float","Numeric","variable","N/A","PARAMETER. Upper bound of the clamp range.",                     "1",        "Must be ≥ minVal.",                               "Param – clamp"],
    ["(return)", "int | float","Numeric","variable","N/A","RETURN VALUE. value clamped to [minVal, maxVal].",               "1.0",      "minVal ≤ return ≤ maxVal.",                       "Return – clamp"],
]
for r in clamp_rows:
    fill = param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill)
    add_row(ws, ROW, r, fill); ROW += 1

style_section(ws, ROW, "   GetGuiAssets()", PatternFill("solid", start_color="4472C4")); ROW += 1
gga_rows = [
    ["sortedAssetList","list","List of GUIBase","variable","N/A","LOCAL. Flat ordered list of GUI objects sorted by zIndex (low to high).",                 "[<GUIBase>, ...]","Populated from sortedAssetDic values.",            "Local – GetGuiAssets"],
    ["sortedAssetDic", "dict","{ int: [GUIBase] }","variable","N/A","LOCAL. Intermediate dict grouping GUI objects by their zIndex before final sort.",    "{1:[...], 2:[...]}","Keys are ints (zIndex); values are lists.",        "Local – GetGuiAssets"],
    ["(return)",       "list","List of GUIBase","variable","N/A","RETURN VALUE. All GUI objects sorted by ascending zIndex for rendering order.",           "[<GUIBase>, ...]","Must be iterable list of GUIBase instances.",       "Return – GetGuiAssets"],
]
for r in gga_rows:
    fill = ret_fill if "RETURN" in r[6] else white_fill
    add_row(ws, ROW, r, fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: FUNCTIONS – TextLabel.py
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  FUNCTIONS — TextLabel.py"); ROW += 1

style_section(ws, ROW, "   GetFont(Name)", PatternFill("solid", start_color="4472C4")); ROW += 1
gf2_rows = [
    ["Name",    "str",                 "Font name string","variable","≤30 chars","PARAMETER. System font name to load or retrieve from cache.",     "\"monospace\"",  "Must be a valid pygame system font name.",          "Param – GetFont"],
    ["(return)","pygame.freetype.SysFont","N/A (object)","variable","N/A",       "RETURN VALUE. Cached or freshly loaded SysFont object for Name.", "<SysFont>",     "Always a valid SysFont; size defaults to 1.",       "Return – GetFont"],
]
for r in gf2_rows:
    fill = param_fill if "PARAM" in r[6] else ret_fill
    add_row(ws, ROW, r, fill); ROW += 1

style_section(ws, ROW, "   GetScaledTextSize(text, font, abSize)", PatternFill("solid", start_color="4472C4")); ROW += 1
gsts_rows = [
    ["text",       "str",   "String",           "variable","N/A",        "PARAMETER. Text to measure. Must not be empty or None.",                          "\"Population:120\"","Raises ValueError if empty or None.",                "Param – GetScaledTextSize"],
    ["font",       "pygame.freetype.SysFont","N/A (object)","variable","N/A","PARAMETER. Font object whose size is temporarily adjusted for measurement.", "<SysFont>",        "Must be a valid SysFont object.",                    "Param – GetScaledTextSize"],
    ["abSize",     "tuple", "(float, float)",   "72",      "(w_px, h_px)","PARAMETER. Absolute pixel dimensions (width, height) the text must fit within.","(200.0, 30.0)",    "Both values must be > 0.",                           "Param – GetScaledTextSize"],
    ["ab_xs",      "float", "Pixels",           "24",      "N/A (local)", "LOCAL. Unpacked pixel width from abSize.",                                       "200.0",            "Derived from abSize[0].",                            "Local – GetScaledTextSize"],
    ["ab_ys",      "float", "Pixels",           "24",      "N/A (local)", "LOCAL. Unpacked pixel height from abSize.",                                      "30.0",             "Derived from abSize[1].",                            "Local – GetScaledTextSize"],
    ["base_size",  "int",   "Integer (pt)",     "28",      "N/A (local)", "LOCAL. Reference font size (100pt) used as measurement baseline.",               "100",              "Fixed constant.",                                    "Local – GetScaledTextSize"],
    ["text_width", "int",   "Pixels",           "28",      "N/A (local)", "LOCAL. Measured pixel width of text at base_size.",                             "180",              "Must be > 0; otherwise returns 1.",                  "Local – GetScaledTextSize"],
    ["text_height","int",   "Pixels",           "28",      "N/A (local)", "LOCAL. Measured pixel height of text at base_size.",                            "24",               "Must be > 0; otherwise returns 1.",                  "Local – GetScaledTextSize"],
    ["scale_x",    "float", "Decimal ratio",    "24",      "N/A (local)", "LOCAL. Horizontal scale factor: ab_xs / text_width.",                           "1.11",             "Positive float.",                                    "Local – GetScaledTextSize"],
    ["scale_y",    "float", "Decimal ratio",    "24",      "N/A (local)", "LOCAL. Vertical scale factor: ab_ys / text_height.",                            "1.25",             "Positive float.",                                    "Local – GetScaledTextSize"],
    ["scale",      "float", "Decimal ratio",    "24",      "N/A (local)", "LOCAL. Minimum of scale_x and scale_y so text fits both dimensions.",           "1.11",             "Must be > 0.",                                       "Local – GetScaledTextSize"],
    ["new_size",   "int",   "Integer (pt)",     "28",      "N/A (local)", "LOCAL. Calculated font size in points that fits the bounding box.",             "111",              "min 1.",                                             "Local – GetScaledTextSize"],
    ["(return)",   "int",   "Integer (pt)",     "28",      "N/A",         "RETURN VALUE. Font size in points that fits text within abSize.",               "111",              "≥ 1.",                                               "Return – GetScaledTextSize"],
]
for r in gsts_rows:
    fill = param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill)
    add_row(ws, ROW, r, fill); ROW += 1

style_section(ws, ROW, "   DetermineWrap(text, font, absoluteSize)", PatternFill("solid", start_color="4472C4")); ROW += 1
dw_rows = [
    ["text",         "str",   "String",          "variable","N/A",        "PARAMETER. Text to wrap. Returns [] if falsy.",                                  "\"Ore In Storage: 1200\"","Non-empty string for meaningful output.",            "Param – DetermineWrap"],
    ["font",         "pygame.freetype.SysFont","N/A (object)","variable","N/A","PARAMETER. Font used for width measurement during word-wrapping.",          "<SysFont>",        "Valid SysFont.",                                     "Param – DetermineWrap"],
    ["absoluteSize", "tuple", "(float, float)",  "72",      "(w_px, h_px)","PARAMETER. Pixel area the wrapped text must fit within.",                       "(200.0, 40.0)",    "Both values > 0.",                                   "Param – DetermineWrap"],
    ["words",        "list",  "List of str",     "variable","N/A",        "LOCAL. Text split on the separator character.",                                  "['Ore','In','Storage:','1200']","Non-empty list.",                      "Local – DetermineWrap"],
    ["rows",         "int",   "Integer",         "28",      "N/A (local)", "LOCAL. Loop counter for the number of rows being attempted.",                   "2",                "1 to len(words).",                                   "Local – DetermineWrap"],
    ["lines",        "list",  "List of str",     "variable","N/A",        "LOCAL. Accumulated lines for the current row-count attempt.",                   "['Ore In','Storage: 1200']","Non-empty list when text is non-empty.",       "Local – DetermineWrap"],
    ["(return)",     "list",  "List of str",     "variable","N/A",        "RETURN VALUE. List of wrapped text lines that fit within absoluteSize.",        "['Ore In','Storage: 1200']","Empty list if text is falsy; otherwise list of str.", "Return – DetermineWrap"],
]
for r in dw_rows:
    fill = param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill)
    add_row(ws, ROW, r, fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: FUNCTIONS – Button.py
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  FUNCTIONS — Button.py"); ROW += 1
style_section(ws, ROW, "   getButtonList()", PatternFill("solid", start_color="4472C4")); ROW += 1
gbl_rows = [
    ["(return)","dict","{ 'MouseUp': [...], 'MouseDown': [...] }","variable","N/A","RETURN VALUE. Module-level buttonList dict mapping event names to callback lists.","{ 'MouseUp':[fn], 'MouseDown':[fn] }","Must contain 'MouseUp' and 'MouseDown' keys.","Return – getButtonList"],
]
add_row(ws, ROW, gbl_rows[0], ret_fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: FUNCTIONS – Image.py
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  FUNCTIONS — Image.py"); ROW += 1
style_section(ws, ROW, "   getImage(ImagePath)", PatternFill("solid", start_color="4472C4")); ROW += 1
gi_rows = [
    ["ImagePath","str",              "File path string","variable","path string","PARAMETER. Relative file path to the image asset to load.",                r"Assets\bg.png",   "Must be a valid path to a pygame-loadable image file.","Param – getImage"],
    ["(return)", "pygame.Surface",  "N/A (object)",    "variable","N/A",        "RETURN VALUE. Loaded (and cached) pygame Surface with alpha channel.",       "<Surface>",        "Valid pygame Surface with convert_alpha() applied.",  "Return – getImage"],
]
for r in gi_rows:
    fill = param_fill if "PARAM" in r[6] else ret_fill
    add_row(ws, ROW, r, fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: FUNCTIONS – main.py (key ones)
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  FUNCTIONS — main.py"); ROW += 1

style_section(ws, ROW, "   AttemptToGetInt(IntString)", PatternFill("solid", start_color="4472C4")); ROW += 1
atgi_rows = [
    ["IntString","str",      "String",   "variable","N/A",  "PARAMETER. A string to attempt conversion to int. Typically from a UI textbox.",       "\"5\"",   "Any string; non-numeric strings result in return value of -1.","Param – AttemptToGetInt"],
    ["a",        "int",      "Integer",  "28",      "N/A",  "LOCAL. Stores result of int() conversion; initialised to -1 as error sentinel.",      "-1",      "Either -1 (error) or the parsed integer.",                    "Local – AttemptToGetInt"],
    ["(return)", "int",      "Integer",  "28",      "N/A",  "RETURN VALUE. Parsed integer if conversion succeeds; -1 if conversion fails.",        "5",       "Any integer, or -1 on failure.",                              "Return – AttemptToGetInt"],
]
for r in atgi_rows:
    fill = param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill)
    add_row(ws, ROW, r, fill); ROW += 1

style_section(ws, ROW, "   UpdateTransactionBalance(_)  — inner: CheckItem(TB, Price, Buying, MaxItemsInTransaction)", PatternFill("solid", start_color="4472C4")); ROW += 1
utb_rows = [
    ["_",                    "any",    "N/A",          "N/A",     "N/A",       "PARAMETER (outer). Ignored argument; required by property-changed signal callback signature.",           "_",              "Accepted but unused.",                                                  "Param – UpdateTransactionBalance"],
    ["RemainingBal",         "int/float","Integer or decimal","variable","N/A","LOCAL (outer & inner via nonlocal). Running balance after applying all simulated transactions.",          "3200",           "Starts as Money; decremented/incremented by CheckItem calls.",          "Local – UpdateTransactionBalance / CheckItem"],
    ["TB",                   "Textbox","N/A (object)", "variable","N/A",       "PARAMETER (inner CheckItem). The UI textbox whose text contains the transaction quantity.",               "<Textbox>",      "Must be a Textbox instance with .Textlabel.Text property.",             "Param – CheckItem (UpdateTransactionBalance)"],
    ["Price",                "int",    "Integer ($)",  "28",      "N/A",       "PARAMETER (inner CheckItem). Unit price of the item being transacted.",                                  "3200",           "Must be > 0 to avoid divide-by-zero in floor() upstream.",              "Param – CheckItem (UpdateTransactionBalance)"],
    ["Buying",               "bool",   "True / False", "28",      "N/A",       "PARAMETER (inner CheckItem). True = purchasing (debit); False = selling (credit).",                     "True",           "Boolean only.",                                                         "Param – CheckItem (UpdateTransactionBalance)"],
    ["MaxItemsInTransaction","int",    "Integer",      "28",      "N/A",       "PARAMETER (inner CheckItem). Upper limit on quantity allowed in this transaction.",                      "5",              "Must be ≥ 0.",                                                          "Param – CheckItem (UpdateTransactionBalance)"],
    ["ItemsInTransaction",   "int",    "Integer",      "28",      "N/A",       "LOCAL (inner). Parsed integer quantity from textbox. Capped at MaxItemsInTransaction.",                  "3",              "≥ 0 after validation; capped at MaxItemsInTransaction.",                "Local – CheckItem (UpdateTransactionBalance)"],
    ["ItemCost",             "int/float","Integer or decimal","variable","N/A","LOCAL (inner). Total cost = ItemsInTransaction × Price.",                                                 "9600",           "Non-negative.",                                                         "Local – CheckItem (UpdateTransactionBalance)"],
    ["(return – CheckItem)", "int",    "Integer",      "28",      "N/A",       "RETURN VALUE of inner CheckItem. Number of items validated for this transaction.",                       "3",              "0 on error or empty input; otherwise validated quantity.",              "Return – CheckItem (UpdateTransactionBalance)"],
    ["(return – outer)",     "None",   "None",         "N/A",     "N/A",       "RETURN VALUE of UpdateTransactionBalance. Updates RemainingBalLabel only; no return value.",            "None",           "Always None.",                                                          "Return – UpdateTransactionBalance"],
]
for r in utb_rows:
    fill = param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill)
    add_row(ws, ROW, r, fill); ROW += 1

style_section(ws, ROW, "   ProcessTransactions()  — inner: CheckItem(TB, Price, Buying, MaxItemsInTransaction)", PatternFill("solid", start_color="4472C4")); ROW += 1
pt_note = [
    ["TB",                   "Textbox","N/A (object)", "variable","N/A","PARAMETER (inner CheckItem). Same role as in UpdateTransactionBalance – the source textbox.",                    "<Textbox>",      "Must be Textbox instance.",                         "Param – CheckItem (ProcessTransactions)"],
    ["Price",                "int",    "Integer ($)",  "28",      "N/A","PARAMETER (inner CheckItem). Unit price – same role as UpdateTransactionBalance version.",                        "3200",           "> 0.",                                              "Param – CheckItem (ProcessTransactions)"],
    ["Buying",               "bool",   "True / False", "28",      "N/A","PARAMETER (inner CheckItem). Same role as UpdateTransactionBalance version.",                                     "False",          "Boolean only.",                                     "Param – CheckItem (ProcessTransactions)"],
    ["MaxItemsInTransaction","int",    "Integer",      "28",      "N/A","PARAMETER (inner CheckItem). Same role as UpdateTransactionBalance version.",                                     "5",              "≥ 0.",                                              "Param – CheckItem (ProcessTransactions)"],
    ["RemainingBal",         "int/float","Integer or decimal","variable","N/A","LOCAL (outer, nonlocal in inner). Running balance; written back to global Money at end.",                "3200",           "Updated by CheckItem calls.",                       "Local – ProcessTransactions / CheckItem"],
    ["foodPurchased",        "int",    "Integer",      "28",      "N/A","LOCAL. Return value of CheckItem for food purchase. Used to update currentSatfication.",                         "40",             "≥ 0.",                                              "Local – ProcessTransactions"],
    ["(return – CheckItem)", "int",    "Integer",      "28",      "N/A","RETURN VALUE of inner CheckItem. Validated transaction quantity (same as UpdateTransactionBalance version).",    "3",              "0 on error; otherwise quantity.",                   "Return – CheckItem (ProcessTransactions)"],
    ["(return – outer)",     "None",   "None",         "N/A",     "N/A","RETURN VALUE of ProcessTransactions. Updates globals directly; no return value.",                               "None",           "Always None.",                                      "Return – ProcessTransactions"],
]
for r in pt_note:
    fill = param_fill if "PARAM" in r[6] else (ret_fill if "RETURN" in r[6] else white_fill)
    add_row(ws, ROW, r, fill); ROW += 1

style_section(ws, ROW, "   UpdateVariables() / Events() / GoToNextTerm() / StartNewGame() / ToggleAllUIVisiblity(Toggle)", PatternFill("solid", start_color="4472C4")); ROW += 1
misc_main = [
    ["Toggle",       "bool",  "True / False", "28",  "N/A","PARAMETER – ToggleAllUIVisiblity. Boolean passed to each UI element's .Visible property.","True","Boolean only.",     "Param – ToggleAllUIVisiblity"],
    ["EventRan",     "float", "Decimal 0–1",  "24",  "N/A","LOCAL – Events(). Random float [0,1) used to determine which random event fires this term.","0.42","0.0 ≤ value < 1.0.", "Local – Events"],
    ["eventOccured", "bool",  "True / False", "28",  "N/A","LOCAL – Events(). Set to True when an event condition is matched; triggers UI display and sleep.","True","Boolean only.", "Local – Events"],
    ["LastFrameTime","float", "Unix timestamp","24",  "N/A","LOCAL – main(). Time of the last rendered frame; used to calculate delta time.",           "1711234567.3","Positive float; set via time.time().", "Local – main"],
    ["ElapedTime",   "float", "Decimal (s)",  "24",  "N/A","LOCAL – main(). Accumulated elapsed time since game start. Note: typo in source.",          "12.5",         "Non-negative.",             "Local – main"],
    ["FPS_CAP",      "int",   "Integer",      "28",  "N/A","LOCAL – main(). Target maximum frames per second.",                                         "60",           "Fixed at 60.",              "Local – main"],
    ["dt",           "float", "Decimal (s)",  "24",  "N/A","LOCAL – main(). Delta time between current and last frame in seconds.",                     "0.016",        "Non-negative float.",       "Local – main"],
    ["new_width",    "int",   "Pixels",       "28",  "N/A","LOCAL – main(). Calculated new window width on VIDEORESIZE event.",                         "1280",         "> 0.",                      "Local – main"],
    ["new_height",   "int",   "Pixels",       "28",  "N/A","LOCAL – main(). Calculated new window height maintaining aspect ratio.",                    "720",          "> 0.",                      "Local – main"],
]
for i, r in enumerate(misc_main):
    fill = param_fill if "PARAM" in r[6] else white_fill
    add_row(ws, ROW, r, fill); ROW += 1

# ───────────────────────────────────────────────────────────────────────────
# SECTION: CLASS ATTRIBUTES
# ───────────────────────────────────────────────────────────────────────────
style_section(ws, ROW, "▶  CLASS ATTRIBUTES"); ROW += 1

classes_data = [
    # Connection
    ["Connection.DisconnectMethod","callable","N/A (function ref)","variable","N/A","Reference to the function that removes this connection's callback from the event's cb dict.","<function>","Must be callable.",                  "Class attr – Connection"],
    ["Connection.Connected",       "bool",    "True / False",     "28",      "N/A","Whether this connection is still active. Set to False by Disconnect().",                  "True",      "Boolean only.",                       "Class attr – Connection"],
    # Event
    ["Event.cbs",          "dict",  "{ UUID: callable }","variable","N/A","Dictionary of registered callbacks keyed by UUID. Values set to None on disconnect.", "{ uuid: fn }","Values callable or None.",                "Class attr – Event"],
    # SuperClass
    ["SuperClass.ClassName","str",  "String",            "variable","N/A","String name of the class, used in property-validation warnings.",                      "\"TextLabel\"","Non-empty string.",                       "Class attr – SuperClass"],
    ["SuperClass.Name",     "str",  "String",            "variable","N/A","Display name; initialised equal to ClassName.",                                        "\"TextLabel\"","Non-empty string.",                       "Class attr – SuperClass"],
    ["SuperClass._Events",  "dict", "{ str: Event }",    "variable","N/A","Maps signal property names to their Event objects. Populated by GetPropertyChangedSignal().", "{ 'Text': <Event> }","Keys must be in SignalProperties.", "Class attr – SuperClass"],
    ["SuperClass.ValidProperties","list","List of str",  "variable","N/A","Allowed property names. Writes to unlisted names generate a warning.",                "['Pos','Size',...]","List of strings.",                   "Class attr – SuperClass"],
    ["SuperClass.SignalProperties","list","List of str", "variable","N/A","Properties for which property-change events can be created.",                         "['Text','Visible']","List of strings.",                   "Class attr – SuperClass"],
    # GUIBase
    ["GUIBase.Pos",          "tuple","(float, float)",   "72",      "(x, y)","Normalised (0–1) position as fraction of screen, centred.",                        "(0.5, 0.5)",  "Both values 0.0–1.0.",                    "Class attr – GUIBase"],
    ["GUIBase.Size",         "tuple","(float, float)",   "72",      "(w, h)","Normalised (0–1) size as fraction of screen.",                                     "(0.1, 0.05)", "Both values > 0.0.",                      "Class attr – GUIBase"],
    ["GUIBase.AbsolutePos",  "tuple","(float, float) px","72",      "pixels","Computed pixel position (top-left corner). Updated each frame in refresh().",      "(350.0, 200.0)","Computed; do not set directly.",        "Class attr – GUIBase"],
    ["GUIBase.AbsoluteSize", "tuple","(float, float) px","72",      "pixels","Computed pixel size. Updated each frame in refresh().",                            "(100.0, 30.0)","Computed; do not set directly.",         "Class attr – GUIBase"],
    ["GUIBase.BackgroundColor","tuple","(R,G,B) ints",   "72",      "(R,G,B)","RGB colour for background rectangle.",                                            "(255,255,255)","Each channel 0–255.",                    "Class attr – GUIBase"],
    ["GUIBase.BackgroundTransparency","float","0.0–1.0", "24",      "decimal","0 = fully opaque, 1 = fully transparent (no rect drawn).",                        "1",            "Clamped to [0,1] by clamp().",            "Class attr – GUIBase"],
    ["GUIBase.zIndex",       "int",  "Integer",          "28",      "N/A","Render order; lower values drawn first (behind higher zIndex elements).",              "1",            "Positive integer.",                       "Class attr – GUIBase"],
    ["GUIBase.Visible",      "bool", "True / False",     "28",      "N/A","Controls whether the element is rendered and receives input.",                        "True",         "Boolean only.",                           "Class attr – GUIBase"],
    ["GUIBase.UIAspectRatio","float | None","Decimal or None","24",  "N/A","If set, overrides height to maintain width:height ratio (width is primary axis).",   "None",         "Positive float or None.",                 "Class attr – GUIBase"],
    # TextLabel
    ["TextLabel.Text",       "str",  "String",           "variable","N/A","The display string rendered onto the label each frame.",                              "\"Population:120\"","Any string; empty string skips render.", "Class attr – TextLabel"],
    ["TextLabel.TextColor",  "tuple","(R,G,B) ints",     "72",      "(R,G,B)","Foreground colour of rendered text.",                                             "(255,255,255)","Each channel 0–255.",                    "Class attr – TextLabel"],
    ["TextLabel.TextFont",   "str",  "Font name string", "variable","N/A","System font name used to render text.",                                               "\"monospace\"","Must be a valid pygame system font name.", "Class attr – TextLabel"],
    ["TextLabel.TextSize",   "int",  "Integer (pt)",     "28",      "N/A","Manual font size; used only when TextScaled is False.",                               "14",           "≥ 1.",                                    "Class attr – TextLabel"],
    ["TextLabel.TextScaled", "bool", "True / False",     "28",      "N/A","If True, font size auto-scales to fit AbsoluteSize.",                                 "True",         "Boolean only.",                           "Class attr – TextLabel"],
    ["TextLabel.TextWrapped","bool", "True / False",     "28",      "N/A","If True and TextScaled, text wraps across multiple lines to fit the bounding box.",   "True",         "Boolean only.",                           "Class attr – TextLabel"],
    # Button
    ["Button.MouseDown",     "Event","N/A (Event obj)",  "variable","N/A","Event fired when a mouse-down click is detected within the button's bounds.",          "<Event>",      "Must be an Event instance.",              "Class attr – Button"],
    ["Button.MouseUp",       "Event","N/A (Event obj)",  "variable","N/A","Event fired when a mouse-up is detected within the button's bounds.",                  "<Event>",      "Must be an Event instance.",              "Class attr – Button"],
    ["Button.MouseClickOff", "Event","N/A (Event obj)",  "variable","N/A","Event fired when a click occurs outside the button's bounds.",                         "<Event>",      "Must be an Event instance.",              "Class attr – Button"],
    # Textbox
    ["Textbox.Textlabel",    "TextLabel","N/A (object)", "variable","N/A","The underlying TextLabel that displays the current textbox content.",                  "<TextLabel>",  "Must be a TextLabel instance.",           "Class attr – Textbox"],
    ["Textbox.Button",       "Button","N/A (object)",   "variable","N/A","The underlying Button that handles click detection for focus.",                        "<Button>",     "Must be a Button instance.",              "Class attr – Textbox"],
    ["Textbox.TypingIn",     "bool", "True / False",    "28",      "N/A","Whether this textbox currently has keyboard focus.",                                   "False",        "Boolean only.",                           "Class attr – Textbox"],
    # Image
    ["Image.ImagePath",      "str",  "File path string","variable","path","Relative path to the image file. Changing this triggers image reload via signal.",    r"Assets\bg.png","Valid relative path to a supported image format.","Class attr – Image"],
    ["Image.Image",          "pygame.Surface","N/A (object)","variable","N/A","Cached pygame Surface for the current ImagePath.",                                "<Surface>",    "Valid pygame Surface.",                   "Class attr – Image"],
]
for i, r in enumerate(classes_data):
    fill = cls_fill if i % 2 == 0 else PatternFill("solid", start_color="F4CCDA")
    add_row(ws, ROW, r, fill); ROW += 1

wb.save("/home/claude/data_dictionary.xlsx")
print("Saved")

